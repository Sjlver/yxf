"""
yxf: Convert from XLSForm to YAML and back.

CLI Usage:
    To convert an XLSForm to a YAML file: `python -m yxf form.xlsx`.
    To convert a YAML file to an XLSForm: `python -m yxf form.yaml`.

Python API Usage:
    from yxf import (
        xlsform_bytes_to_yaml_string,
        yaml_string_to_xlsform_bytes,
        xlsform_bytes_to_markdown_string,
        markdown_string_to_xlsform_bytes
    )
    
    # Convert in memory
    yaml_output = xlsform_bytes_to_yaml_string(xlsx_bytes)
    xlsx_output = yaml_string_to_xlsform_bytes(yaml_string)
"""

import argparse
import collections
import logging
import pathlib
import re
from io import BytesIO

import markdown_it
import markdown_it.tree
import openpyxl
import openpyxl.styles
import openpyxl.utils
import strictyaml

log = logging.getLogger("yxf.__main__")


def _row_to_dict(headers, values):
    row_dict = collections.OrderedDict()
    for h, v in zip(headers, values):
        if v is None or v == "":
            continue
        if h is None:
            raise ValueError(f"Cell with no column header: {v}")
        row_dict[h] = v
    return row_dict


def _convert_sheet(sheet):
    """Convert an openpyxl sheet to a list of OrderedDicts."""
    headers = _get_headers_from_sheet(sheet)
    result = []
    for row in _content_rows_from_sheet(sheet, values_only=True):
        values = _truncate_row(row)
        values = [_stringify_value(v) for v in values]
        row_dict = _row_to_dict(headers, values)
        if row_dict:
            result.append(row_dict)
    return result


def _get_headers_from_sheet(sheet):
    """Get headers from an openpyxl sheet."""
    for row in sheet.iter_rows(values_only=True):
        return [_stringify_value(h) for h in _truncate_row(row)]
    return []


def _content_rows_from_sheet(sheet, **kwargs):
    """Get content rows (excluding header) from an openpyxl sheet."""
    rows_iter = sheet.iter_rows(**kwargs)
    next(rows_iter)
    return rows_iter


def _truncate_row(row):
    """Returns the row without any empty cells at the end."""
    row = list(row)
    while row and row[-1] is None:
        row.pop()
    return row


def _stringify_value(v):
    """Converts a value to string in a way that's meaningful for values read from Excel."""
    return str(v) if v else ""


def _convert_to_sheet(sheet, rows, keys):
    key_set = set(keys)

    for i, key in enumerate(keys):
        sheet.cell(row=1, column=i + 1, value=key)

    next_row = 2
    previous_list_name = rows[0].get("list_name") if rows else None
    for row in rows:
        if row.get("type") == "begin_group":
            next_row += 1

        if row.get("list_name") != previous_list_name:
            previous_list_name = row.get("list_name")
            next_row += 1

        if not all(k in key_set for k in row.keys()):
            missing_key = next(k for k in row.keys() if k not in key_set)
            raise ValueError(
                f'Invalid key "{missing_key}" in row "{row.get("name", "(unnamed)")}". '
                f"Add it to yxf.headers.{sheet.title} in the YAML file."
            )

        for i, key in enumerate(keys):
            if key in row:
                sheet.cell(row=next_row, column=i + 1, value=row[key])

        next_row += 1

    return sheet


def _write_to_xlsform_memory(form):
    """Write form to Excel workbook in memory, return BytesIO."""
    wb = openpyxl.Workbook()
    for sheet_name in form:
        if sheet_name == "yxf":
            continue
        _convert_to_sheet(
            wb.create_sheet(sheet_name),
            form[sheet_name],
            form["yxf"]["headers"][sheet_name],
        )
    wb.remove(wb.active)
    _make_pretty(wb)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _make_pretty(wb):
    """Apply XLSForm-specific styling to workbook."""
    # Styles
    HEADER_STYLE = openpyxl.styles.NamedStyle(name="header")
    HEADER_STYLE.font = openpyxl.styles.Font(bold=True)

    CODE_STYLE = openpyxl.styles.NamedStyle(name="code")
    CODE_STYLE.font = openpyxl.styles.Font(name="Courier New", color="ff19007d")

    NAME_STYLE = openpyxl.styles.NamedStyle(name="name")
    NAME_STYLE.font = openpyxl.styles.Font(name="Courier New", color="ffa13b16")

    COMMENT_STYLE = openpyxl.styles.NamedStyle(name="comment")
    COMMENT_STYLE.font = openpyxl.styles.Font(name="Courier New", color="ff009c5d")

    NOTE_STYLE = openpyxl.styles.NamedStyle(name="note")
    NOTE_STYLE.font = openpyxl.styles.Font(color="ff555555")

    GROUP_COLORS = [
        ["#ff88d7", "#ff6ec1", "#ff54ab", "#ff3795"],
        ["#ff8e72", "#ff745b", "#ff5a43", "#ff3d29"],
        ["#ffab00", "#ff9300", "#ff7a00", "#ff6200"],
        ["#ffd100", "#ffb900", "#eea200", "#d78b00"],
        ["#cbf300", "#b5db00", "#9fc400", "#8bad00"],
        ["#00ff7c", "#00f165", "#00d94d", "#00c233"],
        ["#00ffe4", "#00f7ce", "#00dfb7", "#00c8a1"],
        ["#00ffff", "#00edff", "#00d5ff", "#00bdef"],
        ["#00edff", "#00d4ff", "#00bcff", "#00a5ff"],
        ["#a0cdff", "#8bb5ff", "#769dff", "#6386ff"],
        ["#faafff", "#e397ff", "#cc80ff", "#b668ff"],
        ["#ff96ff", "#ff7eff", "#ff66fa", "#eb4de3"],
        ["#ff88d7", "#ff6ec1", "#ff54ab", "#ff3795"],
    ]

    for sheet in wb:
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.style = HEADER_STYLE
        sheet.freeze_panes = sheet["A2"]

        sheet_headers = _get_headers_from_sheet(sheet)
        comment_column = sheet_headers.index("#") if "#" in sheet_headers else -1
        type_column = sheet_headers.index("type") if "type" in sheet_headers else -1

        # Set column widths
        widths = [[] for _ in sheet_headers]
        for row in _content_rows_from_sheet(sheet):
            for i, cell in enumerate(row):
                if cell.value:
                    width = max(len(w) for w in str(cell.value).splitlines())
                    widths[i].append(width)

        for i, ws in enumerate(widths):
            col_widths = sorted(ws)
            num_rows = len(col_widths)
            percentile_index = num_rows * 3 // 4
            if i == comment_column:
                estimated_width = 2
            elif col_widths:
                estimated_width = col_widths[percentile_index] + 10
            else:
                estimated_width = 10

            if estimated_width <= 60:
                sheet.column_dimensions[
                    openpyxl.utils.get_column_letter(i + 1)
                ].width = estimated_width
            else:
                sheet.column_dimensions[
                    openpyxl.utils.get_column_letter(i + 1)
                ].width = 60
                for row_index, _ in enumerate(sheet):
                    sheet.cell(
                        row=row_index + 1, column=i + 1
                    ).alignment = openpyxl.styles.Alignment(wrap_text=True)

        # Apply specific styles
        code_columns = set(
            ["calculation", "relevant", "constraint", "repeat_count", "instance_name"]
        )
        for row in _content_rows_from_sheet(sheet):
            for i, cell in enumerate(row):
                if sheet_headers[i] in code_columns:
                    cell.style = CODE_STYLE
                elif sheet_headers[i] == "name":
                    cell.style = NAME_STYLE
                elif sheet_headers[i] == "#":
                    cell.style = COMMENT_STYLE
                elif type_column >= 0 and row[type_column].value == "note":
                    cell.style = NOTE_STYLE

        # Highlight groups
        group_number = 0
        nesting_depth = 0
        if type_column >= 0:
            for row in _content_rows_from_sheet(sheet):
                if str(row[type_column].value).startswith("begin_"):
                    if nesting_depth == 0:
                        group_number += 1
                    nesting_depth += 1

                if nesting_depth > 0:
                    group_colors = GROUP_COLORS[group_number % len(GROUP_COLORS)]
                    cell_color = group_colors[
                        nesting_depth - 1
                        if nesting_depth <= len(group_colors)
                        else len(group_colors) - 1
                    ]
                    if comment_column >= 0:
                        row[comment_column].fill = openpyxl.styles.PatternFill(
                            fgColor="ff" + cell_color[1:], fill_type="solid"
                        )

                if str(row[type_column].value).startswith("end_"):
                    nesting_depth -= 1


def _check_existing_output(filename, force):
    if filename.exists() and not force:
        raise ValueError(f"File already exists (use --force to override): {filename}")


def _ensure_yxf_comment(form, name, file_format):
    desired_comment = f"Converted by yxf, from {name}. Edit the {file_format} file instead of the Excel file."

    # Only add comment if there are rows in survey
    if not form["survey"]:
        return
    
    # Check if we have a comment column
    has_comment_column = "#" in form["yxf"]["headers"]["survey"]
    
    first_line = form["survey"][0] if form["survey"] else {}
    
    # Only manipulate comments if the column exists or we need to add it
    if has_comment_column:
        if "#" not in first_line or not first_line.get("#", "").startswith("Converted by yxf,"):
            form["survey"].insert(0, {"#": desired_comment})
        else:
            form["survey"][0]["#"] = desired_comment
    else:
        # Add comment column if not present
        form["yxf"]["headers"]["survey"].insert(0, "#")
        form["survey"].insert(0, {"#": desired_comment})


def _validate_sheet_name(sheet_name, filename, line):
    if sheet_name not in ["survey", "choices", "settings"]:
        raise ValueError(
            f"{filename}:{line}: Invalid sheet name (must be survey, choices, or settings): {sheet_name}"
        )


def _load_workbook_from_wb(wb):
    """Load workbook data from an openpyxl Workbook object."""
    result = collections.OrderedDict()
    headers = collections.OrderedDict()
    for sheet_name in ["survey", "choices", "settings"]:
        if sheet_name in wb:
            result[sheet_name] = _convert_sheet(wb[sheet_name])
            headers[sheet_name] = _get_headers_from_sheet(wb[sheet_name])
            if headers[sheet_name] and headers[sheet_name][0] != "#":
                if "#" in headers[sheet_name]:
                    raise ValueError(
                        f"The comment column must come first in sheet {sheet_name}."
                    )

    if "survey" not in result:
        raise ValueError('An XLSForm must have a "survey" sheet.')

    result["yxf"] = {"headers": headers}
    return result


def _load_workbook_from_bytes(xlsx_bytes):
    """Load workbook data from Excel bytes."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True)
    return _load_workbook_from_wb(wb)


def xlsform_bytes_to_yaml_string(xlsx_bytes: bytes, source_name: str = "input.xlsx") -> str:
    """Convert XLSForm bytes to YAML string."""
    form = _load_workbook_from_bytes(xlsx_bytes)
    _ensure_yxf_comment(form, source_name, "YAML")
    
    # strictyaml cannot serialize empty lists
    # Add a placeholder comment to empty sheets to avoid this issue
    for sheet_name in ["survey", "choices", "settings"]:
        if sheet_name in form and isinstance(form[sheet_name], list) and len(form[sheet_name]) == 0:
            # Add a minimal comment entry to keep the structure valid
            if "#" not in form["yxf"]["headers"].get(sheet_name, []):
                form["yxf"]["headers"][sheet_name].insert(0, "#")
            form[sheet_name].append({"#": f"Empty {sheet_name} sheet"})
    
    return strictyaml.as_document(form).as_yaml()


def xlsform_bytes_to_markdown_string(xlsx_bytes: bytes, source_name: str = "input.xlsx") -> str:
    """Convert XLSForm bytes to Markdown string."""
    form = _load_workbook_from_bytes(xlsx_bytes)
    _ensure_yxf_comment(form, source_name, "Markdown")

    md = []
    for sheet_name in ["survey", "choices", "settings"]:
        if sheet_name not in form:
            continue

        md.append(f"## {sheet_name}")
        md.append("")

        sheet = form[sheet_name]
        headers = form["yxf"]["headers"][sheet_name]
        header_indices = dict(zip(headers, range(len(headers))))

        # Before we render the table, look for comments and render those.
        # We simply put them as paragraphs in the Markdown file.
        for row in sheet:
            if "#" in row:
                if row["#"]:
                    md.append(row["#"])
                    md.append("")
                del row["#"]

        if headers[0] == "#":
            headers.pop(0)
            del header_indices["#"]
            header_indices = {k: v - 1 for (k, v) in header_indices.items()}

        for i, row in enumerate(sheet):
            for k, v in row.items():
                # Markdown does not support multi-line entries in cells. Check
                # and complain if needed.
                if "\n" in v:
                    log.warning(
                        "%s:%d Multi-line value for column %s.\n"
                        "Markdown does not support multi-line values. Use YAML instead.",
                        source_name,
                        i + 2,
                        k,
                    )
                    v = v.replace("\n", " ")
                # Markdown uses "|" as a table cell separator. Escape it if it
                # occurs in one of the values. And duplicate each escape character.
                row[k] = v.replace("\\", "\\\\").replace("|", "\\|")

        # Find column widths
        widths = [len(h) for h in headers]
        for row in sheet:
            for k, v in row.items():
                i = header_indices[k]
                widths[i] = max(widths[i], len(v))

        # Render the table
        header_row = [h.ljust(w) for (h, w) in zip(headers, widths)]
        md.append(f"| {' | '.join(header_row)} |")
        separator_row = ["-" * w for w in widths]
        md.append(f"| {' | '.join(separator_row)} |")
        for row in sheet:
            if not row:
                continue
            formatted_row = [row.get(h, "").ljust(w) for (h, w) in zip(headers, widths)]
            md.append(f"| {' | '.join(formatted_row)} |")
        md.append("")

    return "\n".join(md)


def yaml_string_to_xlsform_bytes(yaml_str: str, source_name: str = "input.yaml") -> bytes:
    """Convert YAML string to XLSForm bytes."""
    form = strictyaml.load(yaml_str).data

    if "yxf" not in form:
        raise ValueError('YAML file must have a "yxf" entry.')
    if "survey" not in form:
        raise ValueError('YAML file must have a "survey" entry.')
    _ensure_yxf_comment(form, source_name, "YAML")
    
    output = _write_to_xlsform_memory(form)
    return output.getvalue()


def markdown_string_to_xlsform_bytes(md_str: str, source_name: str = "input.md") -> bytes:
    """Convert Markdown string to XLSForm bytes."""
    parser = markdown_it.MarkdownIt("js-default")
    ast = markdown_it.tree.SyntaxTreeNode(parser.parse(md_str))
    form = collections.OrderedDict()
    form_headers = collections.OrderedDict()
    sheet_name = None
    for node in ast:
        if node.tag == "h2":
            sheet_name = node.children[0].content
            _validate_sheet_name(sheet_name, source_name, node.map[0])
            result = []
        elif node.tag == "p":
            content = node.children[0].content
            match = re.match(r"%%\s*(.*)", content)
            if match:
                sheet_name = match.group(1)
                _validate_sheet_name(sheet_name, source_name, node.map[0])
            else:
                # Other paragraphs are treated as comments and added to the
                # beginning of the current sheet.
                result.append({"#": content})
        elif node.tag == "table":
            if not sheet_name:
                raise ValueError(
                    f"{source_name}:{node.map[0]}: No sheet name specified for table."
                )
            thead, tbody = node.children
            headers = [c.children[0].content for c in thead.children[0].children]
            add_comment_column = headers[0] != "#" and result and "#" in result[0]
            if add_comment_column:
                headers.insert(0, "#")
            rows = tbody.children
            rows = [[c.children[0].content for c in row.children] for row in rows]
            for values in rows:
                if add_comment_column:
                    values.insert(0, "")
                row_dict = _row_to_dict(headers, values)
                if row_dict:
                    result.append(row_dict)
            form[sheet_name] = result
            form_headers[sheet_name] = headers
    form["yxf"] = {"headers": form_headers}

    _ensure_yxf_comment(form, source_name, "Markdown")
    output = _write_to_xlsform_memory(form)
    return output.getvalue()

def xlsform_to_yaml(filename: pathlib.Path, target: pathlib.Path):
    """Convert XLSForm file `filename` to YAML file `target`."""
    log.info("xlsform_to_yaml: %s -> %s", filename, target)
    
    # Read file into memory
    with open(filename, "rb") as f:
        xlsx_bytes = f.read()
    
    # Process in memory
    yaml_str = xlsform_bytes_to_yaml_string(xlsx_bytes, filename.name)
    
    # Write result to file
    with open(target, "w", encoding="utf-8") as f:
        f.write(yaml_str)


def xlsform_to_markdown(filename: pathlib.Path, target: pathlib.Path):
    """Convert XLSForm file `filename` to Markdown file `target`."""
    log.info("xlsform_to_markdown: %s -> %s", filename, target)
    
    # Read file into memory
    with open(filename, "rb") as f:
        xlsx_bytes = f.read()
    
    # Process in memory
    md_str = xlsform_bytes_to_markdown_string(xlsx_bytes, filename.name)
    
    # Write result to file
    with open(target, "w", encoding="utf-8") as f:
        f.write(md_str)


def yaml_to_xlsform(filename: pathlib.Path, target: pathlib.Path):
    """Convert YAML file `filename` to XLSForm file `target`."""
    log.info("yaml_to_xlsform: %s -> %s", filename, target)
    
    # Read file into memory
    with open(filename, encoding="utf-8") as f:
        yaml_str = f.read()
    
    # Process in memory
    xlsx_bytes = yaml_string_to_xlsform_bytes(yaml_str, filename.name)
    
    # Write result to file
    with open(target, "wb") as f:
        f.write(xlsx_bytes)


def markdown_to_xlsform(filename: pathlib.Path, target: pathlib.Path):
    """Convert Markdown file `filename` to XLSForm file `target`."""
    log.info("markdown_to_xlsform: %s -> %s", filename, target)
    
    # Read file into memory
    with open(filename, encoding="utf-8") as f:
        md_str = f.read()
    
    # Process in memory
    xlsx_bytes = markdown_string_to_xlsform_bytes(md_str, filename.name)
    
    # Write result to file
    with open(target, "wb") as f:
        f.write(xlsx_bytes)


def main():
    """yxf: Convert from XLSForm to YAML and back."""

    logging.basicConfig(level=logging.DEBUG)
    logging.getLogger("markdown_it").setLevel(logging.INFO)

    parser = argparse.ArgumentParser(
        description="Convert from XLSForm to YAML and back"
    )
    parser.add_argument("file", type=pathlib.Path, help="a file to be converted")
    parser.add_argument(
        "--markdown",
        action="store_true",
        help="use Markdown instead of YAML",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=pathlib.Path,
        help="output file name (default: same as input, with extension changed)",
    )
    parser.add_argument(
        "-f",
        "--force",
        action="store_true",
        help="allow overwriting existing output files",
    )
    args = parser.parse_args()

    if args.file.suffix == ".xlsx":
        if args.markdown or (args.output and args.output.suffix == ".md"):
            args.output = args.output or args.file.with_suffix(".md")
            _check_existing_output(args.output, args.force)
            xlsform_to_markdown(args.file, args.output)
        else:
            args.output = args.output or args.file.with_suffix(".yaml")
            _check_existing_output(args.output, args.force)
            xlsform_to_yaml(args.file, args.output)
    elif args.file.suffix == ".yaml":
        args.output = args.output or args.file.with_suffix(".xlsx")
        _check_existing_output(args.output, args.force)
        yaml_to_xlsform(args.file, args.output)
    elif args.file.suffix == ".md":
        args.output = args.output or args.file.with_suffix(".xlsx")
        _check_existing_output(args.output, args.force)
        markdown_to_xlsform(args.file, args.output)
    else:
        raise ValueError(f"Unrecognized file extension: {args.file}")


if __name__ == "__main__":
    main()
