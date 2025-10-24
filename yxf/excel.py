"""Functions for reading and writing XLSForm Excel files.

This module provides the core library functions for converting between XLSForm
Excel files and Python dictionaries. Functions accept file-like objects (BinaryIO)
rather than filenames for maximum flexibility.
"""

import collections
import logging
from typing import BinaryIO

import openpyxl

from . import xlsform

log = logging.getLogger(__name__)


def row_to_dict(headers, values):
    """Convert a row of values to an ordered dictionary using headers as keys.

    Args:
        headers: List of column headers
        values: List of cell values

    Returns:
        OrderedDict with non-empty values mapped to their headers

    Raises:
        ValueError: If a non-empty value has no corresponding header

    >>> row_to_dict(["name", "type", "label"], ["q1", "text", "Question 1"])
    OrderedDict({'name': 'q1', 'type': 'text', 'label': 'Question 1'})
    >>> row_to_dict(["name", "type"], ["q1", ""])
    OrderedDict({'name': 'q1'})
    >>> row_to_dict(["name", "type"], ["q1", None])
    OrderedDict({'name': 'q1'})
    """
    row_dict = collections.OrderedDict()
    for h, v in zip(headers, values):
        if v is None or v == "":
            continue
        if h is None:
            raise ValueError(f"Cell with no column header: {v}")
        row_dict[h] = v
    return row_dict


def _convert_sheet(sheet):
    """Convert an Excel sheet to a list of dictionaries.

    Args:
        sheet: openpyxl Worksheet object

    Returns:
        List of OrderedDicts, one per row (excluding header)
    """
    headers = xlsform.headers(sheet)
    result = []
    for row in xlsform.content_rows(sheet, values_only=True):
        values = xlsform.truncate_row(row)
        values = [xlsform.stringify_value(v) for v in values]
        row_dict = row_to_dict(headers, values)
        if row_dict:
            result.append(row_dict)
    return result


def _convert_to_sheet(sheet, rows, keys):
    """Convert a list of dictionaries to an Excel sheet.

    Args:
        sheet: openpyxl Worksheet object to write to
        rows: List of OrderedDicts representing rows
        keys: List of column headers

    Returns:
        The modified sheet

    Raises:
        ValueError: If a row contains a key not in the headers list
    """
    key_set = set(keys)

    for i, key in enumerate(keys):
        sheet.cell(row=1, column=i + 1, value=key)

    next_row = 2
    previous_list_name = rows[0].get("list_name") if rows else None
    for row in rows:
        if row.get("type") == "begin_group":
            next_row += 1

        if row.get("list_name") != previous_list_name:
            previous_list_name = row.get("list_name")
            next_row += 1

        if not all(k in key_set for k in row.keys()):
            missing_key = next(k for k in row.keys() if k not in key_set)
            raise ValueError(
                f'Invalid key "{missing_key}" in row "{row.get("name", "(unnamed)")}". '
                f"Add it to yxf.headers.{sheet.title} in the YAML file."
            )

        for i, key in enumerate(keys):
            if key in row:
                sheet.cell(row=next_row, column=i + 1, value=row[key])

        next_row += 1

    return sheet


def validate_sheet_name(sheet_name, source_name, line):
    """Validate that a sheet name is one of the allowed XLSForm sheets.

    Args:
        sheet_name: Name to validate
        source_name: Name of source file (for error messages)
        line: Line number (for error messages)

    Raises:
        ValueError: If sheet name is not valid

    >>> validate_sheet_name("survey", "test.yaml", 1)
    >>> validate_sheet_name("choices", "test.yaml", 1)
    >>> validate_sheet_name("settings", "test.yaml", 1)
    >>> validate_sheet_name("invalid", "test.yaml", 1)
    Traceback (most recent call last):
        ...
    ValueError: test.yaml:1: Invalid sheet name (must be survey, choices, or settings): invalid
    """
    if sheet_name not in ["survey", "choices", "settings"]:
        raise ValueError(
            f"{source_name}:{line}: Invalid sheet name (must be survey, choices, or settings): {sheet_name}"
        )


def ensure_yxf_comment(form, name, file_format):
    """Ensure the form has a yxf conversion comment in the first row.

    Args:
        form: Form dictionary to modify
        name: Name of source file
        file_format: Format name (e.g., "YAML", "Markdown")
    """
    desired_comment = f"Converted by yxf, from {name}. Edit the {file_format} file instead of the Excel file."

    first_line = form["survey"][0]
    if "#" not in first_line or not first_line["#"].startswith("Converted by yxf,"):
        form["survey"].insert(0, {"#": desired_comment})
    else:
        form["survey"][0]["#"] = desired_comment

    if "#" not in form["yxf"]["headers"]["survey"]:
        form["yxf"]["headers"]["survey"].insert(0, "#")


def read_xlsform(file_obj: BinaryIO) -> dict:
    """Read an XLSForm from a file-like object.

    Args:
        file_obj: Binary file-like object containing Excel data

    Returns:
        Dictionary with sheet names as keys and lists of row dicts as values.
        Also includes a "yxf" key with metadata including headers.

    Raises:
        ValueError: If the workbook doesn't have a "survey" sheet
    """
    wb = openpyxl.load_workbook(file_obj, read_only=True)
    result = collections.OrderedDict()
    headers = collections.OrderedDict()

    for sheet_name in ["survey", "choices", "settings"]:
        if sheet_name in wb:
            result[sheet_name] = _convert_sheet(wb[sheet_name])
            headers[sheet_name] = xlsform.headers(wb[sheet_name])
            if headers[sheet_name] and headers[sheet_name][0] != "#":
                if "#" in headers[sheet_name]:
                    raise ValueError(
                        f"The comment column must come first in sheet {sheet_name}."
                    )

    if "survey" not in result:
        raise ValueError('An XLSForm must have a "survey" sheet.')

    result["yxf"] = {"headers": headers}
    return result


def write_xlsform(form: dict, file_obj: BinaryIO) -> None:
    """Write a form dictionary to an XLSForm Excel file.

    Args:
        form: Dictionary with sheet names as keys and lists of row dicts as values.
              Must include a "yxf" key with headers metadata.
        file_obj: Binary file-like object to write Excel data to

    Raises:
        ValueError: If a row contains a key not in the sheet's headers
    """
    wb = openpyxl.Workbook()
    for sheet_name in form:
        if sheet_name == "yxf":
            continue
        _convert_to_sheet(
            wb.create_sheet(sheet_name),
            form[sheet_name],
            form["yxf"]["headers"][sheet_name],
        )
    if wb.active is not None:
        wb.remove(wb.active)
    xlsform.make_pretty(wb)
    wb.save(file_obj)
