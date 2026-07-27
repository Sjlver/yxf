"""Functions to add XLSForm-specific logic to openpyxl Worksheets."""

import itertools
import logging
import re

import openpyxl.cell.rich_text
import openpyxl.cell.text
import openpyxl.styles
import openpyxl.utils

log = logging.getLogger(__name__)

# The sheets that yxf knows how to convert. Other sheets (for example the
# documentation sheet in the template from xlsform.org) are ignored.
KNOWN_SHEETS = ["survey", "choices", "settings", "entities", "external_choices"]

# Openpyxl styles for various parts of an XLSForm.
HEADER_STYLE = openpyxl.styles.NamedStyle(name="header")
HEADER_STYLE.font = openpyxl.styles.Font(bold=True)

CODE_STYLE = openpyxl.styles.NamedStyle(name="code")
CODE_STYLE.font = openpyxl.styles.Font(name="Courier New", color="ff19007d")

NAME_STYLE = openpyxl.styles.NamedStyle(name="name")
NAME_STYLE.font = openpyxl.styles.Font(name="Courier New", color="ffa13b16")

COMMENT_STYLE = openpyxl.styles.NamedStyle(name="comment")
COMMENT_STYLE.font = openpyxl.styles.Font(name="Courier New", color="ff009c5d")

NOTE_STYLE = openpyxl.styles.NamedStyle(name="note")
NOTE_STYLE.font = openpyxl.styles.Font(color="ff555555")

# Inline font for HTML tags inside otherwise human-readable cells. Tags are
# dimmed and monospaced so that the actual text stands out.
HTML_TAG_FONT = openpyxl.cell.text.InlineFont(rFont="Courier New", color="ff8c8c8c")

# Matches an HTML tag such as "<b>", "</div>" or '<div style="color:red">'. The
# tag name must start with a letter, so that a comparison like "a < b and c > d"
# is not mistaken for a tag. The group is capturing, so that re.split keeps the
# tags.
HTML_TAG_RE = re.compile(r"(</?[a-zA-Z][^<>]*>)")

# Colors for groups. These are essentially "oklch(0.8 - j*0.07, 0.25, 30*i)".
# Each row has a different hue, and values get darker with increasing column.
GROUP_COLORS = [
    ["#ff88d7", "#ff6ec1", "#ff54ab", "#ff3795"],
    ["#ff8e72", "#ff745b", "#ff5a43", "#ff3d29"],
    ["#ffab00", "#ff9300", "#ff7a00", "#ff6200"],
    ["#ffd100", "#ffb900", "#eea200", "#d78b00"],
    ["#cbf300", "#b5db00", "#9fc400", "#8bad00"],
    ["#00ff7c", "#00f165", "#00d94d", "#00c233"],
    ["#00ffe4", "#00f7ce", "#00dfb7", "#00c8a1"],
    ["#00ffff", "#00edff", "#00d5ff", "#00bdef"],
    ["#00edff", "#00d4ff", "#00bcff", "#00a5ff"],
    ["#a0cdff", "#8bb5ff", "#769dff", "#6386ff"],
    ["#faafff", "#e397ff", "#cc80ff", "#b668ff"],
    ["#ff96ff", "#ff7eff", "#ff66fa", "#eb4de3"],
    ["#ff88d7", "#ff6ec1", "#ff54ab", "#ff3795"],
]


# Structural rows that open or close a group. XLSForm accepts both a space and
# an underscore as the separator, so all eight of "begin group", "begin_group",
# "begin repeat", "begin_repeat" and their "end" counterparts are valid.
#
# The match is anchored on purpose. Without anchoring, the metadata type "end"
# (which records the submission end time) would look like the end of a group,
# which silently corrupts the nesting depth for the rest of the sheet.
GROUP_MARKER_RE = re.compile(r"(begin|end)[\s_]+(group|repeat)$")


def parse_group_marker(type_value):
    """Parses a "type" value that opens or closes a group.

    Returns a ("begin" or "end", "group" or "repeat") tuple, or None if the
    type is not a group marker.

    All valid spellings are recognized:

    >>> parse_group_marker("begin group")
    ('begin', 'group')
    >>> parse_group_marker("begin_group")
    ('begin', 'group')
    >>> parse_group_marker("end repeat")
    ('end', 'repeat')
    >>> parse_group_marker("end_repeat")
    ('end', 'repeat')

    Surrounding whitespace and capitalization are ignored:

    >>> parse_group_marker("  Begin  Group ")
    ('begin', 'group')

    Everything else is not a group marker. In particular, the metadata types
    "start" and "end" merely record submission times; they do not delimit a
    group:

    >>> parse_group_marker("end") is None
    True
    >>> parse_group_marker("start") is None
    True
    >>> parse_group_marker("select_one fruits") is None
    True
    >>> parse_group_marker("") is None
    True
    >>> parse_group_marker(None) is None
    True
    """
    if not isinstance(type_value, str):
        return None
    match = GROUP_MARKER_RE.match(type_value.strip().lower())
    if not match:
        return None
    return match.group(1), match.group(2)


def is_group_begin(type_value):
    """Returns whether the type value opens a group or a repeat.

    >>> is_group_begin("begin_repeat")
    True
    >>> is_group_begin("end_repeat")
    False
    """
    marker = parse_group_marker(type_value)
    return marker is not None and marker[0] == "begin"


def is_group_end(type_value):
    """Returns whether the type value closes a group or a repeat.

    >>> is_group_end("end group")
    True
    >>> is_group_end("end")
    False
    """
    marker = parse_group_marker(type_value)
    return marker is not None and marker[0] == "end"


def nesting_levels(type_values, source_name="<input>"):
    """Returns the group nesting level of each of the given rows.

    Level 0 means that a row is not inside any group. A "begin" row and its
    matching "end" row both report the level of the group that they delimit,
    so that the whole group including its delimiters can be highlighted.

    >>> nesting_levels(["begin_group", "text", "begin_group", "text",
    ...                 "end_group", "end_group", "text"])
    [1, 1, 2, 2, 2, 1, 0]

    The metadata types "start" and "end" do not affect nesting:

    >>> nesting_levels(["start", "end", "begin_group", "text", "end_group"])
    [0, 0, 1, 1, 1]

    Unbalanced forms do not raise; surplus "end" rows are reported at level 0.

    >>> nesting_levels(["end_group", "text"])
    [0, 0]
    """
    levels = []
    depth = 0
    for line, type_value in enumerate(type_values):
        marker = parse_group_marker(type_value)
        if marker is None:
            levels.append(depth)
        elif marker[0] == "begin":
            depth += 1
            levels.append(depth)
        else:
            if depth == 0:
                log.warning(
                    '%s:%d: "%s" without a matching "begin %s".',
                    source_name,
                    line + 2,
                    type_value,
                    marker[1],
                )
            levels.append(depth)
            depth = max(0, depth - 1)

    if depth > 0:
        log.warning("%s: %d group(s) are not closed.", source_name, depth)

    return levels


def split_html_tags(value):
    """Splits a string into a list of alternating text and HTML tag parts.

    The parts always concatenate back to the original string. Parts at odd
    indices are HTML tags; parts at even indices are the text in between, and
    may be empty.

    >>> split_html_tags("Hello <b>world</b>")
    ['Hello ', '<b>', 'world', '</b>', '']
    >>> split_html_tags("no markup here")
    ['no markup here']
    >>> split_html_tags("<br>")
    ['', '<br>', '']
    """
    return HTML_TAG_RE.split(value)


def highlight_html(value):
    """Returns a rich-text version of the value with HTML tags highlighted.

    Returns None if the value contains no HTML tags, in which case the plain
    string should be kept as is.

    The parts that are not tags are plain strings, so that they keep whatever
    font the cell itself uses.

    >>> highlight_html("no markup here") is None
    True
    >>> [str(part) for part in highlight_html("Eat <b>fruit</b>")]
    ['Eat ', '<b>', 'fruit', '</b>']
    """
    parts = split_html_tags(value)
    if len(parts) == 1:
        return None

    rich_text = openpyxl.cell.rich_text.CellRichText()
    for i, part in enumerate(parts):
        if not part:
            continue
        if i % 2 == 1:
            rich_text.append(openpyxl.cell.rich_text.TextBlock(HTML_TAG_FONT, part))
        else:
            rich_text.append(part)
    return rich_text


def ensure_yxf_comment(form, name, file_format):
    """Ensure the form has a yxf conversion comment in the first row.

    Args:
        form: Form dictionary to modify
        name: Name of source file
        file_format: Format name (e.g., "YAML", "Markdown")
    """
    desired_comment = f"Converted by yxf, from {name}. Edit the {file_format} file instead of the Excel file."

    first_line = form["survey"][0] if form["survey"] else {}
    if "#" not in first_line or not first_line["#"].startswith("Converted by yxf,"):
        form["survey"].insert(0, {"#": desired_comment})
    else:
        form["survey"][0]["#"] = desired_comment

    if "#" not in form["yxf"]["headers"]["survey"]:
        form["yxf"]["headers"]["survey"].insert(0, "#")


def truncate_row(row):
    """Returns the row without any empty cells at the end.

    >>> truncate_row([1, 2, 3, None, None, None])
    [1, 2, 3]
    """
    row = list(row)
    while row and row[-1] is None:
        row.pop()
    return row


def stringify_value(v):
    """Converts a value to string in a way that's meaningful for values read from Excel.

    >>> stringify_value("hello")
    'hello'
    >>> stringify_value(42)
    '42'
    >>> stringify_value(None)
    ''
    >>> stringify_value("")
    ''
    """
    return str(v) if v else ""


def headers(sheet):
    """Returns the values of the sheet's header row (i.e., the first row)."""

    for row in sheet.iter_rows(values_only=True):
        return [stringify_value(h) for h in truncate_row(row)]

    # If we get here, the sheet is empty.
    return []


def content_rows(sheet, **kwargs):
    """Returns an iterator over the sheet's content rows.

    These are the rows below the header row)."""

    # islice rather than next(), so that a completely empty sheet yields no
    # rows instead of raising StopIteration.
    return itertools.islice(sheet.iter_rows(**kwargs), 1, None)


def row_to_dict(keys, values):
    """Convert a row of values to a dictionary using headers as keys.

    Args:
        headers: List of column headers
        values: List of cell values

    Returns:
        dict with non-empty values mapped to their headers

    Raises:
        ValueError: If a non-empty value has no corresponding header

    >>> row_to_dict(["name", "type", "label"], ["q1", "text", "Question 1"])
    {'name': 'q1', 'type': 'text', 'label': 'Question 1'}
    >>> row_to_dict(["name", "type"], ["q1", ""])
    {'name': 'q1'}
    >>> row_to_dict(["name", "type"], ["q1", None])
    {'name': 'q1'}
    """
    row_dict = {}
    for h, v in zip(keys, values):
        if v is None or v == "":
            continue
        if h is None:
            raise ValueError(f"Cell with no column header: {v}")
        row_dict[h] = v
    return row_dict


def validate_sheet_name(sheet_name, source_name, line):
    """Validate that a sheet name is one of the allowed XLSForm sheets.

    Args:
        sheet_name: Name to validate
        source_name: Name of source file (for error messages)
        line: Line number (for error messages)

    Raises:
        ValueError: If sheet name is not valid

    >>> validate_sheet_name("survey", "test.yaml", 1)
    >>> validate_sheet_name("choices", "test.yaml", 1)
    >>> validate_sheet_name("settings", "test.yaml", 1)
    >>> validate_sheet_name("entities", "test.yaml", 1)
    >>> validate_sheet_name("invalid", "test.yaml", 1)  # doctest: +ELLIPSIS
    Traceback (most recent call last):
        ...
    ValueError: test.yaml:1: Invalid sheet name (must be one of ...): invalid
    """
    if sheet_name not in KNOWN_SHEETS:
        raise ValueError(
            f"{source_name}:{line}: Invalid sheet name "
            f"(must be one of {', '.join(KNOWN_SHEETS)}): {sheet_name}"
        )


# Columns that hold XPath expressions rather than text meant for respondents.
CODE_COLUMNS = frozenset(
    ["calculation", "relevant", "constraint", "repeat_count", "instance_name"]
)

# Columns that hold identifiers or code, and therefore never contain HTML.
NON_PROSE_COLUMNS = CODE_COLUMNS | frozenset(["#", "name", "type", "list_name"])

# Columns wider than this get wrapped text instead of growing further.
MAX_COLUMN_WIDTH = 60


def _set_column_widths(sheet, sheet_headers, comment_column):
    """Sets each column to the 75th percentile of its content width, plus 10."""

    widths = [[] for _ in sheet_headers]
    for row in content_rows(sheet):
        for i, cell in enumerate(row):
            if cell.value and i < len(widths):
                widths[i].append(max(len(w) for w in str(cell.value).splitlines()))

    for i, column_widths in enumerate(widths):
        column_widths = sorted(column_widths)
        if i == comment_column:
            # The comment column only holds a color, so it can stay narrow.
            estimated_width = 2
        elif column_widths:
            estimated_width = column_widths[len(column_widths) * 3 // 4] + 10
        else:
            estimated_width = 10

        column_letter = openpyxl.utils.get_column_letter(i + 1)
        if estimated_width <= MAX_COLUMN_WIDTH:
            sheet.column_dimensions[column_letter].width = estimated_width
        else:
            sheet.column_dimensions[column_letter].width = MAX_COLUMN_WIDTH
            for row_index, _ in enumerate(sheet):
                sheet.cell(row=row_index + 1, column=i + 1).alignment = (
                    openpyxl.styles.Alignment(wrap_text=True)
                )


def _apply_column_styles(sheet, sheet_headers, type_column):
    """Applies fonts and colors to the columns and rows that yxf knows about."""

    for row in content_rows(sheet):
        for i, cell in enumerate(row):
            if i >= len(sheet_headers):
                continue
            if sheet_headers[i] in CODE_COLUMNS:
                cell.style = CODE_STYLE
            elif sheet_headers[i] == "name":
                cell.style = NAME_STYLE
            elif sheet_headers[i] == "#":
                cell.style = COMMENT_STYLE
            elif type_column >= 0 and row[type_column].value == "note":
                cell.style = NOTE_STYLE


def _highlight_groups(sheet, type_column, comment_column):
    """Colors one cell per row to show the group structure of the sheet.

    Each top-level group gets its own hue, and groups nested inside it get
    progressively darker shades of that hue. Both the "begin" and the "end" row
    of a group are colored, so that a group reads as one block.

    The color goes into the comment column if there is one, since that column
    is narrow and holds no text. Otherwise it goes into the first column.
    """

    highlight_column = comment_column if comment_column >= 0 else 0
    rows = list(content_rows(sheet))
    types = [row[type_column].value for row in rows]

    group_number = 0
    for row, type_value, level in zip(rows, types, nesting_levels(types, sheet.title)):
        if level == 1 and is_group_begin(type_value):
            group_number += 1
        if level > 0:
            group_colors = GROUP_COLORS[group_number % len(GROUP_COLORS)]
            cell_color = group_colors[min(level - 1, len(group_colors) - 1)]
            row[highlight_column].fill = openpyxl.styles.PatternFill(
                fgColor="ff" + cell_color[1:], fill_type="solid"
            )


def _highlight_html(sheet, sheet_headers):
    """Dims and monospaces the HTML tags in cells that hold text.

    Labels and hints often contain markup such as <b> or <div style="...">.
    Styling the tags differently from the text makes it much easier to read the
    question that a respondent will actually see.
    """

    for row in content_rows(sheet):
        for i, cell in enumerate(row):
            if i >= len(sheet_headers) or sheet_headers[i] in NON_PROSE_COLUMNS:
                continue
            if isinstance(cell.value, str):
                rich_text = highlight_html(cell.value)
                if rich_text is not None:
                    cell.value = rich_text


def make_pretty(wb: openpyxl.Workbook):
    """Applies styles to the given workbook to make it prettier.

    This function knows about some XLSForm column names and row types, and
    formats them appropriately. It also adds color to highlight the group
    structure of the file, and highlights HTML markup inside cells.
    """
    for sheet in wb:
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.style = HEADER_STYLE
        sheet.freeze_panes = sheet["A2"]

        sheet_headers = headers(sheet)
        comment_column = sheet_headers.index("#") if "#" in sheet_headers else -1
        type_column = sheet_headers.index("type") if "type" in sheet_headers else -1

        _set_column_widths(sheet, sheet_headers, comment_column)
        _apply_column_styles(sheet, sheet_headers, type_column)
        if type_column >= 0:
            _highlight_groups(sheet, type_column, comment_column)
        # Must come last: the passes above read cell values as strings, which
        # rich text is not.
        _highlight_html(sheet, sheet_headers)
