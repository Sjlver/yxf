"""Tests for the formatting that yxf applies to the XLSForm files it writes.

These build a workbook in memory with write_xlsform and then inspect the
result, since that is what a user of the tool actually gets to see.
"""

import io

import openpyxl
import pytest
from openpyxl.cell.rich_text import CellRichText

from yxf.excel import write_xlsform
from yxf.xlsform import GROUP_COLORS, highlight_html, split_html_tags


def build_sheet(rows, headers=("#", "type", "name", "label")):
    """Writes a survey sheet and returns it for inspection."""
    form = {
        "survey": rows,
        "yxf": {"headers": {"survey": list(headers)}},
    }
    excel_bytes = io.BytesIO()
    write_xlsform(form, excel_bytes)
    excel_bytes.seek(0)
    return openpyxl.load_workbook(excel_bytes, rich_text=True)["survey"]


def layout(sheet, type_column=2):
    """Returns one (type, fill) pair per content row.

    The fill is None for an unfilled cell, and the type is None for a blank
    row, which makes the expected layout easy to read in a test.
    """
    result = []
    for row in sheet.iter_rows(min_row=2):
        fill = row[0].fill.fgColor.rgb if row[0].fill.fill_type else None
        result.append((row[type_column - 1].value, fill))
    return result


def color(group_number, level):
    """The fill that a row at the given group and nesting level should have."""
    group_colors = GROUP_COLORS[group_number % len(GROUP_COLORS)]
    return "ff" + group_colors[min(level - 1, len(group_colors) - 1)][1:]


class TestGroupHighlighting:
    """Tests for the colors that show the group structure."""

    def test_group_rows_including_delimiters_are_filled(self):
        sheet = build_sheet(
            [
                {"type": "text", "name": "before"},
                {"type": "begin_group", "name": "g"},
                {"type": "text", "name": "inside"},
                {"type": "end_group", "name": "g"},
                {"type": "text", "name": "after"},
            ]
        )
        assert layout(sheet) == [
            ("text", None),
            (None, None),
            ("begin_group", color(1, 1)),
            ("text", color(1, 1)),
            ("end_group", color(1, 1)),
            (None, None),
            ("text", None),
        ]

    def test_nested_groups_get_darker_shades_of_one_hue(self):
        sheet = build_sheet(
            [
                {"type": "begin_group", "name": "outer"},
                {"type": "begin_group", "name": "inner"},
                {"type": "text", "name": "q"},
                {"type": "end_group", "name": "inner"},
                {"type": "end_group", "name": "outer"},
            ]
        )
        # The blank rows that frame the inner group sit inside the outer one,
        # and so carry the outer group's color.
        assert layout(sheet) == [
            ("begin_group", color(1, 1)),
            (None, color(1, 1)),
            ("begin_group", color(1, 2)),
            ("text", color(1, 2)),
            ("end_group", color(1, 2)),
            (None, color(1, 1)),
            ("end_group", color(1, 1)),
        ]

    def test_each_top_level_group_gets_its_own_hue(self):
        sheet = build_sheet(
            [
                {"type": "begin_group", "name": "a"},
                {"type": "end_group", "name": "a"},
                {"type": "begin_group", "name": "b"},
                {"type": "end_group", "name": "b"},
            ]
        )
        fills = [fill for _, fill in layout(sheet) if fill]
        assert fills == [color(1, 1), color(1, 1), color(2, 1), color(2, 1)]

    def test_metadata_end_row_does_not_close_the_group(self):
        """Regression test for "type: end" being treated as the end of a group.

        Before this was fixed, the metadata row reset the nesting depth, and
        every row after it lost its highlight.
        """
        sheet = build_sheet(
            [
                {"type": "begin_group", "name": "g"},
                {"type": "start", "name": "start"},
                {"type": "end", "name": "end"},
                {"type": "text", "name": "q"},
                {"type": "end_group", "name": "g"},
            ]
        )
        assert all(fill == color(1, 1) for _, fill in layout(sheet))

    def test_metadata_end_before_any_group(self):
        """A top-level "end" used to drive the depth negative for good."""
        sheet = build_sheet(
            [
                {"type": "start", "name": "start"},
                {"type": "end", "name": "end"},
                {"type": "begin_group", "name": "g"},
                {"type": "text", "name": "q"},
                {"type": "end_group", "name": "g"},
            ]
        )
        assert layout(sheet) == [
            ("start", None),
            ("end", None),
            (None, None),
            ("begin_group", color(1, 1)),
            ("text", color(1, 1)),
            ("end_group", color(1, 1)),
        ]

    @pytest.mark.parametrize(
        "begin,end",
        [
            ("begin group", "end group"),
            ("begin_group", "end_group"),
            ("begin repeat", "end repeat"),
            ("begin_repeat", "end_repeat"),
            ("begin group", "end_group"),
            ("begin_repeat", "end repeat"),
        ],
    )
    def test_every_spelling_is_highlighted(self, begin, end):
        sheet = build_sheet(
            [
                {"type": begin, "name": "g"},
                {"type": "text", "name": "q"},
                {"type": end, "name": "g"},
            ]
        )
        assert all(fill == color(1, 1) for _, fill in layout(sheet))

    def test_fill_goes_in_the_first_column_without_a_comment_column(self):
        """Without a "#" column the fill used to land on the last column."""
        sheet = build_sheet(
            [
                {"type": "begin_group", "name": "g"},
                {"type": "text", "name": "q"},
                {"type": "end_group", "name": "g"},
            ],
            headers=("type", "name", "label"),
        )
        for row in sheet.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            assert row[0].fill.fill_type == "solid", "first column is filled"
            assert row[-1].fill.fill_type is None, "last column is not filled"

    def test_sheet_without_a_type_column_is_not_highlighted(self):
        sheet = build_sheet(
            [{"list_name": "fruits", "name": "apple", "label": "Apple"}],
            headers=("list_name", "name", "label"),
        )
        assert all(cell.fill.fill_type is None for row in sheet for cell in row)


class TestBlankRows:
    """Tests for the blank rows that separate the parts of a sheet."""

    def test_blank_row_before_a_begin_and_after_an_end(self):
        sheet = build_sheet(
            [
                {"type": "text", "name": "before"},
                {"type": "begin_group", "name": "g"},
                {"type": "text", "name": "inside"},
                {"type": "end_group", "name": "g"},
                {"type": "text", "name": "after"},
            ]
        )
        assert [t for t, _ in layout(sheet)] == [
            "text",
            None,
            "begin_group",
            "text",
            "end_group",
            None,
            "text",
        ]

    @pytest.mark.parametrize(
        "begin,end",
        [
            ("begin group", "end group"),
            ("begin_group", "end_group"),
            ("begin repeat", "end repeat"),
            ("begin_repeat", "end_repeat"),
        ],
    )
    def test_repeats_and_all_spellings_get_blank_rows(self, begin, end):
        """Only "begin group" and "begin_group" used to be separated."""
        sheet = build_sheet(
            [
                {"type": "text", "name": "before"},
                {"type": begin, "name": "g"},
                {"type": end, "name": "g"},
                {"type": "text", "name": "after"},
            ]
        )
        assert [t for t, _ in layout(sheet)] == [
            "text",
            None,
            begin,
            end,
            None,
            "text",
        ]

    def test_no_blank_row_at_the_top_of_a_sheet(self):
        sheet = build_sheet(
            [
                {"type": "begin_group", "name": "g"},
                {"type": "end_group", "name": "g"},
            ]
        )
        assert sheet["B2"].value == "begin_group"

    def test_consecutive_blank_rows_are_collapsed(self):
        """An end immediately followed by a begin gets a single blank row."""
        sheet = build_sheet(
            [
                {"type": "begin_group", "name": "a"},
                {"type": "end_group", "name": "a"},
                {"type": "begin_group", "name": "b"},
                {"type": "end_group", "name": "b"},
            ]
        )
        assert [t for t, _ in layout(sheet)] == [
            "begin_group",
            "end_group",
            None,
            "begin_group",
            "end_group",
        ]

    def test_no_trailing_blank_row(self):
        sheet = build_sheet(
            [
                {"type": "begin_group", "name": "g"},
                {"type": "end_group", "name": "g"},
            ]
        )
        assert sheet.max_row == 3

    def test_choice_lists_are_separated(self):
        sheet = build_sheet(
            [
                {"list_name": "fruits", "name": "apple", "label": "Apple"},
                {"list_name": "fruits", "name": "pear", "label": "Pear"},
                {"list_name": "berries", "name": "fig", "label": "Fig"},
            ],
            headers=("list_name", "name", "label"),
        )
        assert [row[0].value for row in sheet.iter_rows(min_row=2)] == [
            "fruits",
            "fruits",
            None,
            "berries",
        ]

    def test_blank_rows_disappear_on_reading(self):
        """Blank rows are cosmetic and must not survive a round trip."""
        from yxf.excel import read_xlsform

        rows = [
            {"type": "begin_group", "name": "g"},
            {"type": "text", "name": "q"},
            {"type": "end_group", "name": "g"},
        ]
        form = {"survey": rows, "yxf": {"headers": {"survey": ["type", "name"]}}}
        excel_bytes = io.BytesIO()
        write_xlsform(form, excel_bytes)
        excel_bytes.seek(0)
        assert read_xlsform(excel_bytes)["survey"] == [
            {"type": "begin_group", "name": "g"},
            {"type": "text", "name": "q"},
            {"type": "end_group", "name": "g"},
        ]


class TestHtmlHighlighting:
    """Tests for highlighting HTML markup inside cells."""

    def test_split_keeps_the_whole_string(self):
        value = '<div style="font-size:18px;">Eat <b>fruit</b><br></div>'
        assert "".join(split_html_tags(value)) == value

    def test_plain_text_is_left_alone(self):
        assert highlight_html("Which fruit do you like?") is None

    def test_comparison_operators_are_not_tags(self):
        """A constraint like "a < b and c > d" must not look like markup."""
        assert highlight_html("Answer must be < 10 and > 0") is None

    def test_tags_are_monospaced_and_dimmed(self):
        rich_text = highlight_html("Eat <b>fruit</b> daily")
        assert isinstance(rich_text, CellRichText)
        tags = [str(p) for p in rich_text if getattr(p, "font", None)]
        prose = [str(p) for p in rich_text if not getattr(p, "font", None)]
        assert tags == ["<b>", "</b>"]
        assert prose == ["Eat ", "fruit", " daily"]
        for part in rich_text:
            if getattr(part, "font", None):
                assert part.font.rFont == "Courier New"
                assert part.font.color.rgb != "FF000000"

    def test_labels_in_a_written_sheet_become_rich_text(self):
        sheet = build_sheet(
            [
                {"type": "note", "name": "n", "label": "Eat <b>fruit</b>"},
                {"type": "text", "name": "q", "label": "Plain label"},
            ]
        )
        assert isinstance(sheet["D2"].value, CellRichText)
        assert isinstance(sheet["D3"].value, str)

    def test_code_columns_are_left_as_plain_text(self):
        """Expressions may contain "<" and must not be touched."""
        sheet = build_sheet(
            [{"type": "integer", "name": "q", "constraint": ". < 10 and . > 0"}],
            headers=("type", "name", "constraint"),
        )
        assert sheet["C2"].value == ". < 10 and . > 0"

    def test_html_survives_a_round_trip_as_plain_text(self):
        from yxf.excel import read_xlsform

        label = '<div style="font-size:18px;">Eat <b>fruit</b><br></div>'
        form = {
            "survey": [{"type": "note", "name": "n", "label": label}],
            "yxf": {"headers": {"survey": ["type", "name", "label"]}},
        }
        excel_bytes = io.BytesIO()
        write_xlsform(form, excel_bytes)
        excel_bytes.seek(0)
        assert read_xlsform(excel_bytes)["survey"][0]["label"] == label
