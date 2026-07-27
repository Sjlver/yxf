"""Unit tests for Excel-specific logic."""

import io

import pytest
import openpyxl

from yxf.excel import read_xlsform, write_xlsform


class TestReadXlsform:
    """Tests for read_xlsform function."""

    def test_missing_survey_sheet_raises_error(self):
        """Test that missing survey sheet raises ValueError."""
        # Create a workbook without a survey sheet (but with valid content)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "choices"
        sheet.append(["list_name", "name", "label"])
        sheet.append(["colors", "red", "Red"])
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        with pytest.raises(ValueError, match='must have a "survey" sheet'):
            read_xlsform(excel_bytes)

    def test_comment_column_not_first_raises_error(self):
        """Test that comment column not in first position raises error."""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "survey"
        sheet.append(["type", "#", "name"])  # Comment column not first
        sheet.append(["text", "A comment", "q1"])
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        with pytest.raises(ValueError, match="comment column must come first"):
            read_xlsform(excel_bytes)


class TestWriteXlsform:
    """Tests for write_xlsform function."""

    def test_invalid_key_in_row_raises_error(self):
        """Test that a row with an invalid key raises ValueError."""
        form = {
            "survey": [{"name": "q1", "invalid_key": "value"}],
            "yxf": {"headers": {"survey": ["name", "type"]}},
        }

        excel_bytes = io.BytesIO()
        with pytest.raises(ValueError, match='Invalid key "invalid_key"'):
            write_xlsform(form, excel_bytes)

    def test_minimal_form_writes_successfully(self):
        """Test that a minimal valid form can be written."""
        form = {
            "survey": [{"name": "q1", "type": "text", "label": "Question 1"}],
            "yxf": {"headers": {"survey": ["name", "type", "label"]}},
        }

        excel_bytes = io.BytesIO()
        write_xlsform(form, excel_bytes)
        excel_bytes.seek(0)

        # Verify it can be read back
        form_read = read_xlsform(excel_bytes)
        assert "survey" in form_read
        assert len(form_read["survey"]) == 1
        assert form_read["survey"][0]["name"] == "q1"

    def test_empty_sheet_does_not_raise(self):
        """Test that a sheet with headers but no rows can be written and read."""
        form = {
            "survey": [{"name": "q1", "type": "text"}],
            "choices": [],
            "yxf": {
                "headers": {
                    "survey": ["name", "type"],
                    "choices": ["list_name", "name", "label"],
                }
            },
        }

        excel_bytes = io.BytesIO()
        write_xlsform(form, excel_bytes)
        excel_bytes.seek(0)

        form_read = read_xlsform(excel_bytes)
        assert form_read["choices"] == []
        assert form_read["yxf"]["headers"]["choices"] == ["list_name", "name", "label"]


class TestExtraSheets:
    """Tests for the sheets beyond survey, choices and settings."""

    @pytest.mark.parametrize("sheet_name", ["entities", "external_choices"])
    def test_extra_sheet_survives_a_roundtrip(self, sheet_name):
        """These sheets used to be dropped silently on every conversion."""
        form = {
            "survey": [{"type": "text", "name": "q1"}],
            sheet_name: [{"list_name": "fruits", "label": "Apple"}],
            "yxf": {
                "headers": {
                    "survey": ["type", "name"],
                    sheet_name: ["list_name", "label"],
                }
            },
        }

        excel_bytes = io.BytesIO()
        write_xlsform(form, excel_bytes)
        excel_bytes.seek(0)

        form_read = read_xlsform(excel_bytes)
        assert form_read[sheet_name] == [{"list_name": "fruits", "label": "Apple"}]

    def test_unknown_sheets_are_ignored(self):
        """A documentation sheet must not end up in the converted form."""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "survey"
        sheet.append(["type", "name"])
        sheet.append(["text", "q1"])
        notes = wb.create_sheet("Start here")
        notes.append(["Some documentation that is not part of the form"])
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        form_read = read_xlsform(excel_bytes)
        assert "Start here" not in form_read
        assert set(form_read) == {"survey", "yxf"}
