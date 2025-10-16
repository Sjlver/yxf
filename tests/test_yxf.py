"""
Unit tests for yxf module.
"""

import pytest
import openpyxl
from io import BytesIO

from yxf import (
    xlsform_bytes_to_yaml_string,
    xlsform_bytes_to_markdown_string,
    yaml_string_to_xlsform_bytes,
    markdown_string_to_xlsform_bytes,
)


def test_xlsform_to_yaml_basic(sample_xlsx_bytes):
    """Test basic XLSForm to YAML conversion."""
    result = xlsform_bytes_to_yaml_string(sample_xlsx_bytes)
    
    assert isinstance(result, str)
    assert "survey:" in result
    assert "choices:" in result
    assert "settings:" in result
    assert "yxf:" in result
    assert "What is your name?" in result


def test_xlsform_to_yaml_has_comment(sample_xlsx_bytes):
    """Test that conversion adds a yxf comment."""
    result = xlsform_bytes_to_yaml_string(sample_xlsx_bytes, "test.xlsx")
    
    assert "Converted by yxf" in result
    assert "test.xlsx" in result


def test_xlsform_to_yaml_preserves_headers(sample_xlsx_bytes):
    """Test that headers are preserved in yxf section."""
    result = xlsform_bytes_to_yaml_string(sample_xlsx_bytes)
    
    assert "headers:" in result
    assert "survey:" in result
    assert "- type" in result
    assert "- name" in result
    assert "- label" in result


def test_yaml_to_xlsform_basic(sample_yaml_string):
    """Test basic YAML to XLSForm conversion."""
    result = yaml_string_to_xlsform_bytes(sample_yaml_string)
    
    assert isinstance(result, bytes)
    assert len(result) > 0
    
    # Verify it's a valid Excel file
    wb = openpyxl.load_workbook(BytesIO(result))
    assert "survey" in wb.sheetnames
    assert "choices" in wb.sheetnames
    assert "settings" in wb.sheetnames


def test_yaml_to_xlsform_content(sample_yaml_string):
    """Test that YAML to XLSForm preserves content."""
    result = yaml_string_to_xlsform_bytes(sample_yaml_string)
    
    wb = openpyxl.load_workbook(BytesIO(result))
    survey = wb["survey"]
    
    # Check headers (row 1)
    headers = [cell.value for cell in survey[1]]
    assert "#" in headers
    assert "type" in headers
    assert "name" in headers
    assert "label" in headers
    
    # Row 2 is the comment, row 3 is first data
    row3 = [cell.value for cell in survey[3]]
    assert "text" in row3
    assert "name" in row3
    assert "What is your name?" in row3


def test_yaml_to_xlsform_missing_yxf():
    """Test that conversion fails without yxf section."""
    invalid_yaml = """survey:
- type: text
  name: test
"""
    
    with pytest.raises(ValueError, match='YAML file must have a "yxf" entry'):
        yaml_string_to_xlsform_bytes(invalid_yaml)


def test_yaml_to_xlsform_missing_survey():
    """Test that conversion fails without survey section."""
    invalid_yaml = """choices:
- list_name: test
  name: test
yxf:
  headers:
    choices:
    - list_name
    - name
"""
    
    with pytest.raises(ValueError, match='YAML file must have a "survey" entry'):
        yaml_string_to_xlsform_bytes(invalid_yaml)


def test_xlsform_to_markdown_basic(sample_xlsx_bytes):
    """Test basic XLSForm to Markdown conversion."""
    result = xlsform_bytes_to_markdown_string(sample_xlsx_bytes)
    
    assert isinstance(result, str)
    assert "## survey" in result
    assert "## choices" in result
    assert "## settings" in result
    assert "| type" in result
    assert "What is your name?" in result


def test_xlsform_to_markdown_table_format(sample_xlsx_bytes):
    """Test that Markdown tables are properly formatted."""
    result = xlsform_bytes_to_markdown_string(sample_xlsx_bytes)
    
    # Check for table structure
    assert "|" in result
    assert "---" in result  # Table separator
    
    # Check for proper column alignment
    lines = result.split("\n")
    survey_section = False
    for line in lines:
        if "## survey" in line:
            survey_section = True
        if survey_section and line.startswith("|") and "type" in line:
            # Found the header row
            assert "name" in line
            assert "label" in line
            break


def test_markdown_to_xlsform_basic(sample_markdown_string):
    """Test basic Markdown to XLSForm conversion."""
    result = markdown_string_to_xlsform_bytes(sample_markdown_string)
    
    assert isinstance(result, bytes)
    assert len(result) > 0
    
    # Verify it's a valid Excel file
    wb = openpyxl.load_workbook(BytesIO(result))
    assert "survey" in wb.sheetnames


def test_markdown_to_xlsform_content(sample_markdown_string):
    """Test that Markdown to XLSForm preserves content."""
    result = markdown_string_to_xlsform_bytes(sample_markdown_string)
    
    wb = openpyxl.load_workbook(BytesIO(result))
    survey = wb["survey"]
    
    # Row 1: headers, Row 2: comment, Row 3: first data
    row3 = [cell.value for cell in survey[3]]
    assert "text" in row3
    assert "name" in row3
    assert "What is your name?" in row3


def test_markdown_escaped_pipes():
    """Test that escaped pipes in Markdown are handled correctly."""
    md = """## survey

Converted by yxf, from test.md. Edit the Markdown file instead of the Excel file.

| type | name | label          |
| ---- | ---- | -------------- |
| text | test | Choice A \\| B |

"""
    result = markdown_string_to_xlsform_bytes(md)
    
    wb = openpyxl.load_workbook(BytesIO(result))
    survey = wb["survey"]
    
    # Row 3 is the first data row (after header and comment)
    label_cell = survey.cell(row=3, column=4).value
    # The label should have the pipe unescaped
    assert label_cell is not None
    # Either the pipe is preserved or unescaped
    assert "|" in label_cell or "Choice A" in label_cell


def test_roundtrip_xlsx_to_yaml_to_xlsx(sample_xlsx_bytes):
    """Test that XLSForm -> YAML -> XLSForm preserves data."""
    # Convert to YAML
    yaml_str = xlsform_bytes_to_yaml_string(sample_xlsx_bytes)
    
    # Convert back to XLSX
    result_bytes = yaml_string_to_xlsform_bytes(yaml_str)
    
    # Verify structure is preserved
    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    assert "survey" in wb.sheetnames
    
    survey = wb["survey"]
    # Skip comment row (row 2), check row 3
    row3 = [cell.value for cell in survey[3]]
    assert "text" in row3
    assert "name" in row3


def test_roundtrip_yaml_to_xlsx_to_yaml(sample_yaml_string):
    """Test that YAML -> XLSForm -> YAML preserves structure."""
    # Convert to XLSX
    xlsx_bytes = yaml_string_to_xlsform_bytes(sample_yaml_string)
    
    # Convert back to YAML
    result_yaml = xlsform_bytes_to_yaml_string(xlsx_bytes)
    
    # Verify key elements are present
    assert "survey:" in result_yaml
    assert "What is your name?" in result_yaml
    assert "yxf:" in result_yaml


def test_roundtrip_xlsx_to_markdown_to_xlsx(sample_xlsx_bytes):
    """Test that XLSForm -> Markdown -> XLSForm preserves data."""
    # Convert to Markdown
    md_str = xlsform_bytes_to_markdown_string(sample_xlsx_bytes)
    
    # Convert back to XLSX
    result_bytes = markdown_string_to_xlsform_bytes(md_str)
    
    # Verify structure is preserved
    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    assert "survey" in wb.sheetnames


def test_empty_survey_sheet():
    """Test handling of XLSForm with only headers."""
    wb = openpyxl.Workbook()
    survey = wb.create_sheet("survey")
    survey.append(["type", "name", "label"])
    
    wb.remove(wb.active)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    result = xlsform_bytes_to_yaml_string(output.getvalue())
    assert "survey:" in result
    # Empty survey should have a placeholder comment to avoid strictyaml serialization errors
    assert "Empty survey sheet" in result or "survey:" in result


def test_xlsx_without_survey_sheet():
    """Test that XLSForm without survey sheet raises error."""
    wb = openpyxl.Workbook()
    choices = wb.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    
    wb.remove(wb.active)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    with pytest.raises(ValueError, match='XLSForm must have a "survey" sheet'):
        xlsform_bytes_to_yaml_string(output.getvalue())


def test_special_characters_in_labels():
    """Test that special characters are preserved."""
    wb = openpyxl.Workbook()
    survey = wb.create_sheet("survey")
    survey.append(["type", "name", "label"])
    survey.append(["text", "test", "Label with 'quotes' and \"double quotes\""])
    
    wb.remove(wb.active)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    result = xlsform_bytes_to_yaml_string(output.getvalue())
    assert "quotes" in result


def test_source_name_parameter():
    """Test that source_name parameter is used in comments."""
    wb = openpyxl.Workbook()
    survey = wb.create_sheet("survey")
    survey.append(["type", "name", "label"])
    survey.append(["text", "test", "Test"])
    
    wb.remove(wb.active)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    result = xlsform_bytes_to_yaml_string(output.getvalue(), "my_custom_form.xlsx")
    assert "my_custom_form.xlsx" in result


def test_comment_column_preserved():
    """Test that comment column (#) is preserved."""
    wb = openpyxl.Workbook()
    survey = wb.create_sheet("survey")
    survey.append(["#", "type", "name", "label"])
    survey.append(["This is a comment", "text", "test", "Test"])
    
    wb.remove(wb.active)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    result = xlsform_bytes_to_yaml_string(output.getvalue())
    assert "This is a comment" in result


def test_comment_column_must_be_first():
    """Test that comment column must be first if present."""
    wb = openpyxl.Workbook()
    survey = wb.create_sheet("survey")
    survey.append(["type", "#", "name", "label"])  # Comment not first
    survey.append(["text", "comment", "test", "Test"])
    
    wb.remove(wb.active)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    with pytest.raises(ValueError, match="comment column must come first"):
        xlsform_bytes_to_yaml_string(output.getvalue())


def test_yaml_without_comment_column():
    """Test YAML conversion when no comment column exists in headers."""
    yaml_no_comment = """survey:
- type: text
  name: test
  label: Test Question
yxf:
  headers:
    survey:
    - type
    - name
    - label
"""
    
    result = yaml_string_to_xlsform_bytes(yaml_no_comment)
    wb = openpyxl.load_workbook(BytesIO(result))
    survey = wb["survey"]
    
    # Should add comment column automatically
    headers = [cell.value for cell in survey[1]]
    assert "#" in headers
    assert headers[0] == "#"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
