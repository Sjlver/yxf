import pytest
import openpyxl
from io import BytesIO


@pytest.fixture
def sample_xlsx_bytes():
    """Create a minimal valid XLSForm as bytes."""
    wb = openpyxl.Workbook()
    
    # Create survey sheet
    survey = wb.create_sheet("survey")
    survey.append(["type", "name", "label"])
    survey.append(["text", "name", "What is your name?"])
    survey.append(["integer", "age", "What is your age?"])
    
    # Create choices sheet
    choices = wb.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    choices.append(["yes_no", "yes", "Yes"])
    choices.append(["yes_no", "no", "No"])
    
    # Create settings sheet
    settings = wb.create_sheet("settings")
    settings.append(["form_title", "form_id"])
    settings.append(["Sample Form", "sample_form"])
    
    wb.remove(wb.active)  # Remove default sheet
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def sample_yaml_string():
    """Create a minimal valid YAML string."""
    return """survey:
- '#': Converted by yxf, from test.yaml. Edit the YAML file instead of the Excel
    file.
- type: text
  name: name
  label: What is your name?
- type: integer
  name: age
  label: What is your age?
choices:
- list_name: yes_no
  name: 'yes'
  label: 'Yes'
- list_name: yes_no
  name: 'no'
  label: 'No'
settings:
- form_title: Sample Form
  form_id: sample_form
yxf:
  headers:
    survey:
    - '#'
    - type
    - name
    - label
    choices:
    - list_name
    - name
    - label
    settings:
    - form_title
    - form_id
"""


@pytest.fixture
def sample_markdown_string():
    """Create a minimal valid Markdown string."""
    return """## survey

Converted by yxf, from test.md. Edit the Markdown file instead of the Excel file.

| type    | name | label                |
| ------- | ---- | -------------------- |
| text    | name | What is your name?   |
| integer | age  | What is your age?    |

## choices

| list_name | name | label |
| --------- | ---- | ----- |
| yes_no    | yes  | Yes   |
| yes_no    | no   | No    |

## settings

| form_title  | form_id     |
| ----------- | ----------- |
| Sample Form | sample_form |
"""
